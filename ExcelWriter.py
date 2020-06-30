"""
Utility Script to write data to a excel sheet.
"""
import cx_Oracle
import openpyxl.worksheet.worksheet as worksheet
from openpyxl import utils
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

import main

num = 1


def write_and_format(sheet, cursor):
    """
    One call func to do it all
    Parameters
    ----------
    sheet
    cursor

    Returns
    -------
    """
    global num
    write_cursor_to_sheet(sheet, cursor)
    format_sheet(sheet)
    to_table_format(sheet, num)
    num += 1


def write_cursor_to_sheet(sheet: worksheet.Worksheet, cursor: cx_Oracle.Cursor):
    """
    Description: Dumps a prepared cursor into a Excel sheet.
    Parameters
    ----------
    sheet: openpyxl worksheet A sheet to dump it into
    cursor: cx_Oracle cursor prepared with the data
    """

    header= [description[0] for description in cursor.description]
    header.append('Customer No')
    header.append('Customer Name')
    sheet.append(header)
    for row in cursor:
        customer_details = main.get_customers(row[-1])
        lst = list(row)
        lst.append(customer_details[0])
        lst.append(customer_details[1])
        sheet.append(lst)


def format_sheet(sheet: worksheet.Worksheet):
    """
    Description: Format's a Excel sheet so that everything is centered and fits the column's
    Parameters
    ----------
    sheet : worksheet.Worksheet sheet to format
    """

    for column_cells in sheet.columns:
        for cell in column_cells:
            cell.alignment = Alignment('center')
        length = max(len(str(cell.value)) for cell in column_cells)
        length, height = split_sizes(length)
        sheet.column_dimensions[utils.get_column_letter(column_cells[0].column)].width = length + 12


def to_table_format(sheet: worksheet.Worksheet, num):
    """
    Description: Format's a Excel sheet into a table format.
    Parameters
    ----------
    sheet: worksheet.Worksheet A excel sheet
    """

    end_col_char = ord('A') + sheet.max_column - 1
    end_col_char = chr(end_col_char)
    end_row_char = sheet.max_row
    tab = Table(displayName=f"Table{num}", ref=f"A1:{end_col_char}{end_row_char}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True,
                           showColumnStripes=False)

    tab.tableStyleInfo = style
    sheet.add_table(tab)
    sheet.sheet_view.showGridLines = False


def split_sizes(n, height=1):
    if n > 50:
        split_sizes(n // 2, height + 1)
    return n, height
