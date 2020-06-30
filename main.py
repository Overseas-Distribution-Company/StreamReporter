import os
import re
from datetime import datetime
from datetime import date
import cx_Oracle
import pyodbc
from openpyxl import Workbook
import ExcelWriter
from new_shortages import new_shortages
from overseasmail import sendmail


def purge(dir, pattern):
    for f in os.listdir(dir):
        if re.search(pattern, f):
            os.remove(os.path.join(dir, f))


def get_customers(document_no):
    connect2 = pyodbc.connect(f"DRIVER=SQL Server;SERVER=10.0.0.30;"
                              f";DATABASE={'Overseas_LIVE'}; UID=sa; PWD=SQLsrv4fr")
    cursor_l = connect2.cursor()
    cursor_l.execute(
        f'''
        SELECT sl.[Destination No_] AS nr, cu.Name AS name
        
        FROM [ODC$Posted Whse_ Shipment Line] sl
        
        INNER JOIN [ODC$Customer] cu ON
            sl.[Destination No_] = cu.[No_]
            
        WHERE sl.[No_] = '{document_no}'
        ''')

    ret_str_no = ''
    ret_str_name = ''
    found_no = []
    for row in cursor_l:
        if row.nr not in found_no:

            if ret_str_no != '':
                ret_str_no += ', '

            if ret_str_name != '':
                ret_str_name += ', '

            found_no.append(row.nr)
            ret_str_no += row.nr
            ret_str_name += row.name

    return ret_str_no, ret_str_name


# Set-up
connection = cx_Oracle.connect(
    "overseas4read",  # UID
    "OverseasSSW",  # PW
    "OTC-VS-018/ORAMULT"  # Server, SID implied Port ->1521
)
cursor = connection.cursor()

if __name__ == '__main__':
    purge(os.getcwd(), '(.*)xlsx')
    wb = Workbook()

    cursor.execute(
        '''
    SELECT nc.INTERNNUMMER,
           nc.OPDRACHTGEVER,
           nc.DOSSIERNUMMER_ERP,
           nc.DOCUMENT_REFERENTIENUMMER_MRN,
           TO_CHAR(TO_DATE(nc.AANGIFTEDATUM, 'yyyymmdd'), 'dd/mm/yyyy') AS AANGIFTEDATUM,
           SUBSTR(nc.DOSSIERNUMMER_ERP,0, INSTR(nc.DOSSIERNUMMER_ERP, '-')-1) AS REF
    FROM NCTS.NCAANGH0 nc
    WHERE ACTIEFBEDRIJF = 'ODC'
      AND nc.STATUSMESSAGE = 'REL_TRA'
    ORDER BY nc.AANGIFTEDATUM
        '''
    )

    ws = wb.active
    ws.title = 'NCTS-OPEN'
    ExcelWriter.write_and_format(ws, cursor)

    cursor.execute('''
    SELECT d.MESSAGESTATUS,
           d.DECLARATIONID,
           d.COMMERCIALREFERENCE,
           d.ARC_AADREFERENCECODE,
           TO_CHAR(TO_DATE(d.ISSUEDATE, 'yyyymmdd'), 'dd/mm/yyyy') AS ISSUEDATE,
           SUBSTR(d.COMMERCIALREFERENCE,0, INSTR(d.COMMERCIALREFERENCE, '-')-1) AS REF
    FROM PLDA.ESDECLARATION d
    
    WHERE ACTIVECOMPANY = 'ODC'
      AND d.MESSAGETYPE = '1'
      AND d.MESSAGESTATUS = 'ARC_ALL'
    ORDER BY ISSUEDATE DESC
    
    ''')
    ws = wb.create_sheet('EMCS-DEPARTURE OPEN')
    ExcelWriter.write_and_format(ws, cursor)

    cursor.execute('''
    SELECT d.MESSAGESTATUS,
           d.DECLARATIONID,
           d.COMPLEMENTARYINFO,
           d.COMMERCIALREFERENCE,
           d.ARC_AADREFERENCECODE,
           TO_CHAR(TO_DATE(d.ISSUEDATE, 'yyyymmdd'), 'dd/mm/yyyy') AS ISSUEDATE,
           cl.SEQUENCE,
           cl.COMMERCIALDESC,
           cl.QUANTITY,
           cl.SHORTAGEOREXCESS,
           cl.OBSERVEDSHORTAGEEXCESS,
           SUBSTR(d.COMMERCIALREFERENCE,0, INSTR(d.COMMERCIALREFERENCE, '-')-1) AS REF
    FROM PLDA.ESDECLARATION d
    
    LEFT JOIN PLDA.ESDECLARTARIF cl
    ON d.DECLARATIONID = cl.DECLARATIONID
    WHERE ACTIVECOMPANY = 'ODC'
      AND d.MESSAGETYPE = '1'
      AND d.MESSAGESTATUS = 'WRT_NOT' AND d.GLOBALCONCLUSIONRECEIPT = '2'
      AND cl.SHORTAGEOREXCESS <> ' '
    
    ORDER BY ISSUEDATE DESC
    ''')

    ws = wb.create_sheet('EMCS-DEPARTURE TEKORTEN')
    ExcelWriter.write_and_format(ws, cursor)
    new_shortages(ws)

    cursor.execute('''
    SELECT  cp.DECLARATION_ID,
            cp.PRINCIPAL,
            cp.COMMERCIALREFERENCE,
            cp.CUSTOMSMAINREFERENCENUMBER,
            TO_CHAR(TO_DATE(cp.ISSUEDATE, 'yyyymmdd'), 'dd/mm/yyyy') AS ISSUEDATE,
            SUBSTR(cp.COMMERCIALREFERENCE,0, INSTR(cp.COMMERCIALREFERENCE, '-')-1) AS REF
            FROM PLDA.CPDECLARATION cp
            WHERE cp.ACTIVECOMPANY = 'ODC' AND cp.REGIMETYPE = 'A' AND cp.STATUSNUMBER_MESSAGE = 'REL_TRA'
    '''
                   )
    ws = wb.create_sheet('PLDA-EXPORT OPEN')
    ExcelWriter.write_and_format(ws, cursor)

    # Finish Up
    str_time = datetime.now().strftime("%d_%m_%y")
    report_name = f'StreamReport{str_time}.xlsx'
    wb.save(report_name)
    connection.close()

    mailer = sendmail.OverseasMail()
    mailer.sender = 'StreamReport@overseas.be'
    mailer.subject = f'Stream Report {date.today().strftime("%d %B %Y")}'
    mailer.add_receiver('manueldemaerel@overseas.be')
    mailer.add_receiver('rheirman@overseas.be')
    # mailer.add_receiver('orderprocessing@overseas.be')
    # mailer.add_receiver('Declaration@overseas.be')
    # mailer.add_receiver('PRubenska@overseas.be')
    mailer.add_attachment(report_name)
    content = """\
    Beste
    
    In bijlage het stream rapport van deze week.
    
    Automatische mail, do not reply.
    Vragen? Contacteer rheirman@overseas.be
    
    mvg
    Reporting services Overseas
    """
    mailer.message = content
    mailer.send_mail()
